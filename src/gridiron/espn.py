"""External projection baselines (ESPN and friends).

Finding #2 named the real target: naive persistence is the floor, but the
*consensus* is what a projection model actually has to beat, because human
analysts encode offseason information the box scores never contain.

The nflverse route to consensus (`load_ff_rankings`, FantasyPros via
DynastyProcess) is unreachable from some networks, and ESPN's projections
page renders client-side so it cannot be scraped from HTML. So this module
takes the pragmatic path: parse whatever you can copy out of a projections
page, in whatever shape it lands.

Two comparisons are possible and they answer different questions:

- **Current season.** Where do we disagree? Neither side has outcomes yet,
  so this identifies where to think harder, not who is right.
- **Historical seasons.** Which is more accurate? This needs a source's
  *past* projections alongside known results, and is the only version that
  can support a "beats consensus" claim.
"""

from __future__ import annotations

import re
from pathlib import Path

import polars as pl
from scipy.stats import spearmanr

POSITIONS = ("QB", "RB", "WR", "TE", "K", "DST", "D/ST")

# ESPN renders a player cell as e.g. "Ja'Marr Chase Cin WR" or
# "Bijan Robinson Atl RB, FLEX". Team and position trail the name.
_PLAYER_CELL = re.compile(
    r"^(?P<name>.+?)\s+(?P<team>[A-Za-z]{2,3})\s+"
    r"(?P<pos>QB|RB|WR|TE|K|DST|D/ST)\b",
    re.IGNORECASE,
)
_NUMBER = re.compile(r"-?\d+(?:\.\d+)?")

TEAM_FIXUPS = {
    "WSH": "WAS", "JAC": "JAX", "LAR": "LA", "OAK": "LV", "SD": "LAC",
    "STL": "LA", "ARZ": "ARI", "BLT": "BAL", "CLV": "CLE", "HST": "HOU",
}


def parse_pasted(text: str, points_index: int = -1) -> pl.DataFrame:
    """Parse projections copied out of a projections page.

    Deliberately forgiving: copied tables arrive as tabs, runs of spaces,
    or one field per line depending on the browser and how the selection
    was made. Rather than demand a format, this looks for a player cell
    and takes a number from the same line.

    ``points_index`` selects which number on the line is the projection
    (-1 = last, which is where season totals usually sit).
    """
    rows = []
    for raw in text.splitlines():
        line = raw.replace("\t", "  ").strip()
        if not line or len(line) < 4:
            continue
        m = _PLAYER_CELL.match(line)
        if not m:
            continue
        numbers = _NUMBER.findall(line[m.end():])
        if not numbers:
            continue
        try:
            points = float(numbers[points_index])
        except (IndexError, ValueError):
            continue
        team = m.group("team").upper()
        rows.append({
            "player_display_name": m.group("name").strip(),
            "position": m.group("pos").upper().replace("D/ST", "DST"),
            "team": TEAM_FIXUPS.get(team, team),
            "external_points": points,
        })

    if not rows:
        raise ValueError(
            "no projections parsed — paste rows that look like "
            "'Player Name Team POS  123.4'"
        )
    return pl.DataFrame(rows).unique(subset=["player_display_name"], keep="first")


# ESPN's printable cheat sheet is laid out in several columns, so a single
# text line interleaves entries from different position groups. Each entry
# is self-describing though:  "15. (30) Josh Jacobs, GB $27 11"
#   positional rank / overall rank / name / team / auction value / bye
_CHEAT_ENTRY = re.compile(
    r"(?P<pos_rank>\d+)\.\s*\((?P<overall>\d+)\)\s*"
    r"(?P<name>[A-Za-z.'\- ]+?),\s*(?P<team>[A-Z]{2,3})\s*"
    r"\$(?P<value>\d+)\s+(?P<bye>\d+)"
)


def parse_cheat_sheet(source: str | Path) -> pl.DataFrame:
    """Parse an ESPN draft-kit cheat sheet (PDF or extracted text).

    Cheat sheets carry **ranks and auction values rather than projected
    points**, which suits this project: finding #2 established that rank is
    the metric that matters, since you draft an ordering. It also sidesteps
    scoring-format mismatches — point totals differ between formats, draft
    order much less so.

    Position is deliberately *not* taken from the column headers. The
    multi-column layout interleaves position groups on every text line, so
    inferring position from layout is fragile; it is recovered by joining
    to our own projections, which already know each player's position.
    """
    text = _read_text(source)
    rows = []
    for m in _CHEAT_ENTRY.finditer(text):
        rows.append({
            "player_display_name": m.group("name").strip(),
            "team": TEAM_FIXUPS.get(m.group("team").upper(), m.group("team").upper()),
            "external_rank": int(m.group("overall")),
            "auction_value": float(m.group("value")),
            "bye": int(m.group("bye")),
        })
    if not rows:
        raise ValueError("no cheat-sheet entries found")
    return (
        pl.DataFrame(rows)
        .unique(subset=["player_display_name"], keep="first")
        .sort("external_rank")
    )


_RANK_ALIASES = {
    "rank": "external_rank", "overall": "external_rank", "ovr": "external_rank",
    "player": "player_display_name", "name": "player_display_name",
    "pos": "position", "position": "position",
    "team": "team", "tm": "team",
    "bye": "bye",
}


def parse_rank_table(path: str | Path, source: str = "external") -> pl.DataFrame:
    """Load a ranking table (xlsx or csv) that has ranks but no point totals.

    RotoWire, ESPN's cheat sheet and most published "top 250" boards share
    this shape: an overall rank, a player, a position, a team. That is
    enough for the comparison that matters, since finding #2 established
    rank as the metric a drafter actually consumes.

    Unlike the PDF path, these files carry an explicit position column, so
    it is used directly rather than recovered from our own projections.
    """
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        header = [str(c or "").strip() for c in rows[0]]
        df = pl.DataFrame(
            {h: [r[i] for r in rows[1:]] for i, h in enumerate(header) if h},
            strict=False,
        )
    else:
        df = pl.read_csv(path)

    rename = {c: _RANK_ALIASES[c.lower().strip()]
              for c in df.columns if c.lower().strip() in _RANK_ALIASES}
    df = df.rename(rename)

    missing = {"player_display_name", "external_rank"} - set(df.columns)
    if missing:
        raise ValueError(
            f"ranking table is missing required columns: {sorted(missing)}"
        )

    keep = [c for c in ("player_display_name", "position", "team",
                        "external_rank", "bye") if c in df.columns]
    df = df.select(keep).drop_nulls("player_display_name")
    df = df.with_columns(pl.col("external_rank").cast(pl.Int64, strict=False))
    if "team" in df.columns:
        df = df.with_columns(pl.col("team").str.to_uppercase().replace(TEAM_FIXUPS))
    return (
        df.drop_nulls("external_rank")
        .unique(subset=["player_display_name"], keep="first")
        .sort("external_rank")
        .with_columns(pl.lit(source).alias("source"))
    )


def _looks_like_path(source: str | Path) -> bool:
    """Distinguish a filename from a blob of pasted text.

    Calling ``Path(text).exists()`` on a multi-line paste raises
    ``OSError: File name too long`` rather than returning False, so the
    check has to happen before touching the filesystem.
    """
    if isinstance(source, Path):
        return True
    return "\n" not in source and len(source) < 4096


def _read_text(source: str | Path) -> str:
    if not _looks_like_path(source):
        return str(source)
    path = Path(source)
    if path.suffix.lower() == ".pdf":
        import pdfplumber

        with pdfplumber.open(path) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    try:
        if path.exists():
            return path.read_text(encoding="utf-8")
    except OSError:
        pass
    return str(source)


def load_external(path: str | Path, source: str = "espn") -> pl.DataFrame:
    """Load a baseline from CSV (preferred) or pasted text."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        df = pl.read_csv(path)
        rename = {}
        for col in df.columns:
            low = col.lower().strip()
            if low in {"player", "name", "player name"}:
                rename[col] = "player_display_name"
            elif low in {"pos", "position"}:
                rename[col] = "position"
            elif low in {"team", "tm"}:
                rename[col] = "team"
            elif low in {"pts", "points", "proj", "projection", "fpts", "total"}:
                rename[col] = "external_points"
        df = df.rename(rename)
        missing = {"player_display_name", "external_points"} - set(df.columns)
        if missing:
            raise ValueError(f"CSV is missing required columns: {sorted(missing)}")
    else:
        df = parse_pasted(path.read_text(encoding="utf-8"))
    return df.with_columns(pl.lit(source).alias("source"))


def _normalize_name(col: str) -> pl.Expr:
    """Names differ across sources: punctuation, suffixes, casing.

    Matching on a normalised key recovers most of the join that an exact
    string match would silently drop — and a silently dropped join looks
    exactly like 'the model has no opinion on this player'.
    """
    return (
        pl.col(col)
        .str.to_lowercase()
        .str.replace_all(r"\b(jr|sr|ii|iii|iv|v)\.?$", "")
        .str.replace_all(r"[^a-z ]", "")
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
    )


def compare(
    ours: pl.DataFrame,
    external: pl.DataFrame,
    our_points: str = "projected_points",
) -> pl.DataFrame:
    """Join our projections to an external source and rank both.

    Ranks are computed *within position*, because that is the comparison a
    drafter makes and because pooled ranks are dominated by the fact that
    quarterbacks outscore everyone.
    """
    a = ours.with_columns(_normalize_name("player_display_name").alias("_key"))
    b = external.with_columns(_normalize_name("player_display_name").alias("_key"))

    carry = [c for c in ("external_points", "external_rank", "auction_value", "bye")
             if c in b.columns]
    joined = a.join(b.select(["_key"] + carry), on="_key", how="inner").drop("_key")

    # A cheat sheet supplies an overall rank directly; a projections table
    # supplies points we rank ourselves. Either way ranks are recomputed
    # *within position*, because that is the comparison a drafter makes.
    if "external_rank" in joined.columns:
        their = pl.col("external_rank").rank().over("position")
    else:
        their = pl.col("external_points").rank(descending=True).over("position")

    out = joined.with_columns(
        pl.col(our_points).rank(descending=True).over("position").alias("our_rank"),
        their.alias("their_rank"),
    ).with_columns(
        # Positive = we rank the player higher than they do.
        (pl.col("their_rank") - pl.col("our_rank")).alias("rank_delta"),
    )
    # A cheat sheet has no point totals to difference against.
    if "external_points" in out.columns:
        out = out.with_columns(
            (pl.col(our_points) - pl.col("external_points")).alias("points_delta")
        )
    return out.sort("rank_delta")


_AGREEMENT_SCHEMA = {
    "position": pl.String,
    "n": pl.Int64,
    "spearman": pl.Float64,
    "mean_abs_rank_delta": pl.Float64,
}


def agreement(
    comparison: pl.DataFrame,
    by_position: bool = True,
    min_players: int = 3,
) -> pl.DataFrame:
    """Rank correlation between the two sources, per position.

    Returns an empty but correctly-typed frame when nothing clears
    ``min_players``, rather than raising from a schema-less DataFrame —
    an unhelpful error at the end of a long pipeline is worse than an
    obviously empty result.
    """
    rows = []
    frames = (
        comparison.partition_by("position", as_dict=True).items()
        if by_position else [(("ALL",), comparison)]
    )
    for key, sub in frames:
        if sub.height < min_players:
            continue
        rho = spearmanr(
            sub["our_rank"].to_numpy(), sub["their_rank"].to_numpy()
        ).statistic
        rows.append({
            "position": key[0] if isinstance(key, tuple) else key,
            "n": sub.height,
            "spearman": float(rho),
            "mean_abs_rank_delta": float(sub["rank_delta"].abs().mean()),
        })
    if not rows:
        return pl.DataFrame(schema=_AGREEMENT_SCHEMA)
    return pl.DataFrame(rows, schema=_AGREEMENT_SCHEMA).sort("position")


def biggest_disagreements(comparison: pl.DataFrame, n: int = 15) -> pl.DataFrame:
    """Players the two sources rank most differently — where to think hardest."""
    cols = [
        "player_display_name", "position", "team",
        "projected_points", "external_points", "our_rank", "their_rank",
        "rank_delta",
    ]
    have = [c for c in cols if c in comparison.columns]
    top = comparison.sort("rank_delta").head(n).select(have)
    bottom = comparison.sort("rank_delta", descending=True).head(n).select(have)
    return pl.concat([top, bottom]).unique(subset=["player_display_name"], keep="first")
